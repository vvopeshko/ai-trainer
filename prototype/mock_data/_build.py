"""
Парсер workouts_2026.xlsx → workouts.json для прототипа мини-аппа.

Структура xlsx:
    Sheets (по месяцам): Дата, Тренировка, Упражнение, Подход 1..6, Всего подходов
    Формат ячейки подхода: "reps×weight" (например, "10×60") или "reps" (bodyweight).

Выход JSON:
    {
      "workouts": [                    # хронологический список Workout
        {
          "id": "w-20260102-1",
          "date": "2026-01-02",
          "name": "2 Chest Back Shoulders B",
          "programSlug": "jan-split",  # псевдо-программа на основе префикса имени
          "exercises": [
            {
              "id": "machine-chest-press",
              "name": "Machine Chest Press",
              "bodyweight": false,
              "sets": [
                {"reps": 10, "weightKg": 60},
                {"reps": 10, "weightKg": 70},
                {"reps": 5, "weightKg": 70}
              ]
            },
            ...
          ]
        },
        ...
      ],
      "exercises": [                   # каталог уникальных упражнений со статистикой
        {
          "id": "machine-chest-press",
          "name": "Machine Chest Press",
          "bodyweight": false,
          "totalSessions": 7,
          "totalSets": 21,
          "maxWeightKg": 70,
          "lastUsed": "2026-04-15"
        },
        ...
      ],
      "programs": [                    # обнаруженные "псевдо-программы"
        {
          "slug": "jan-split",
          "label": "Январь: 2-дневный сплит",
          "firstSeen": "2026-01-02",
          "lastSeen": "2026-01-30",
          "workoutNames": ["2 Chest Back Shoulders B", "2 Legs + Arms B", ...]
        },
        ...
      ],
      "stats": {                       # месячная агрегация для UI прогресса
        "2026-01": {
          "workouts": 8, "exerciseSessions": 63, "totalSets": 181,
          "totalTonnageKg": 38290, "uniqueExercises": 27
        },
        ...
      }
    }
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

# Путь к исходному xlsx — можно перезадать через env MOCK_SOURCE.
# По умолчанию ищем рядом со скриптом: prototype/mock_data/workouts_2026.xlsx
import os

SCRIPT_DIR = Path(__file__).resolve().parent
SRC = Path(os.environ.get("MOCK_SOURCE", SCRIPT_DIR / "workouts_2026.xlsx"))
DST = SCRIPT_DIR / "workouts.json"

if not SRC.exists():
    raise SystemExit(
        f"Не найден xlsx: {SRC}\n"
        f"Скопируй свой файл сюда как 'workouts_2026.xlsx' либо укажи путь:\n"
        f"  MOCK_SOURCE=/path/to/file.xlsx python3 _build.py"
    )

SET_RE = re.compile(r"^\s*(\d+)\s*[×x]\s*(\d+(?:[.,]\d+)?)\s*$")


def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def parse_cell(raw) -> dict | None:
    """Вернёт {'reps': int, 'weightKg': float} либо {'reps': int} (bodyweight)."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = SET_RE.match(s)
    if m:
        reps = int(m.group(1))
        weight = float(m.group(2).replace(",", "."))
        return {"reps": reps, "weightKg": weight}
    # просто число повторов (bodyweight)
    if s.isdigit():
        return {"reps": int(s)}
    # иногда бывает что-то нестандартное (например "8-10") — сохраняем как заметку
    return {"reps": 0, "note": s}


def month_program(date_str: str, workout_name: str) -> tuple[str, str]:
    """Псевдо-программа: по месяцу + префиксу имени."""
    month = date_str[:7]
    # в январе у него "2 Chest ...", "2 Legs ...", "2 Upper Body A"
    # с февраля — ".1 Chest Triceps", ".2 Back Biceps", ".3 Legs", ".4 Shoulders Arms"
    if month == "2026-01":
        return "jan-split", "Январь: 2-дневный сплит"
    else:
        return "push-pull-legs-arms", "PPL+Arms (4-дневный сплит)"


def main():
    wb = load_workbook(SRC, data_only=True)

    # Группа "Workout" = (date, workoutName). Внутри — строки упражнений.
    # Используем dict, чтобы агрегировать строки с тем же date+workoutName.
    workouts_map: dict[tuple[str, str], dict] = {}
    workout_order: list[tuple[str, str]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Инструкция":
            continue
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            date_cell, workout_name, exercise_name, *set_cells_raw = row[:9]
            if not workout_name or not exercise_name:
                continue
            date_str = date_cell.date().isoformat() if hasattr(date_cell, "date") else str(date_cell)
            workout_name = str(workout_name).strip()
            exercise_name = str(exercise_name).strip()

            set_cells = set_cells_raw[:6]  # "Подход 1..6"
            parsed_sets = [p for p in (parse_cell(c) for c in set_cells) if p is not None]
            if not parsed_sets:
                continue

            # Bodyweight = ни в одном подходе нет weightKg
            bodyweight = all("weightKg" not in s for s in parsed_sets)

            key = (date_str, workout_name)
            if key not in workouts_map:
                workouts_map[key] = {
                    "id": f"w-{date_str.replace('-', '')}-{len(workout_order) + 1}",
                    "date": date_str,
                    "name": workout_name,
                    "exercises": [],
                }
                workout_order.append(key)
                slug, label = month_program(date_str, workout_name)
                workouts_map[key]["programSlug"] = slug
                workouts_map[key]["programLabel"] = label

            workouts_map[key]["exercises"].append(
                {
                    "id": slugify(exercise_name),
                    "name": exercise_name,
                    "bodyweight": bodyweight,
                    "sets": parsed_sets,
                }
            )

    workouts = [workouts_map[k] for k in workout_order]
    # Сортируем хронологически
    workouts.sort(key=lambda w: (w["date"], w["name"]))

    # ─── Каталог упражнений со статистикой ───
    ex_stats: dict[str, dict] = {}
    for w in workouts:
        for ex in w["exercises"]:
            key = ex["id"]
            if key not in ex_stats:
                ex_stats[key] = {
                    "id": ex["id"],
                    "name": ex["name"],
                    "bodyweight": ex["bodyweight"],
                    "totalSessions": 0,
                    "totalSets": 0,
                    "maxWeightKg": 0,
                    "firstUsed": w["date"],
                    "lastUsed": w["date"],
                }
            stats = ex_stats[key]
            stats["totalSessions"] += 1
            stats["totalSets"] += len(ex["sets"])
            stats["lastUsed"] = max(stats["lastUsed"], w["date"])
            stats["firstUsed"] = min(stats["firstUsed"], w["date"])
            for s in ex["sets"]:
                w_kg = s.get("weightKg", 0) or 0
                if w_kg > stats["maxWeightKg"]:
                    stats["maxWeightKg"] = w_kg

    exercises = sorted(ex_stats.values(), key=lambda e: -e["totalSets"])

    # ─── Программы ───
    progs: dict[str, dict] = {}
    for w in workouts:
        slug = w["programSlug"]
        if slug not in progs:
            progs[slug] = {
                "slug": slug,
                "label": w["programLabel"],
                "firstSeen": w["date"],
                "lastSeen": w["date"],
                "workoutNames": set(),
            }
        progs[slug]["lastSeen"] = max(progs[slug]["lastSeen"], w["date"])
        progs[slug]["firstSeen"] = min(progs[slug]["firstSeen"], w["date"])
        progs[slug]["workoutNames"].add(w["name"])

    programs = []
    for p in progs.values():
        programs.append(
            {
                **p,
                "workoutNames": sorted(p["workoutNames"]),
            }
        )
    programs.sort(key=lambda p: p["firstSeen"])

    # ─── Статистика по месяцам ───
    stats_by_month: dict[str, dict] = defaultdict(
        lambda: {
            "workouts": 0,
            "exerciseSessions": 0,
            "totalSets": 0,
            "totalTonnageKg": 0.0,
            "uniqueExercises": set(),
        }
    )
    for w in workouts:
        month = w["date"][:7]
        s = stats_by_month[month]
        s["workouts"] += 1
        for ex in w["exercises"]:
            s["exerciseSessions"] += 1
            s["uniqueExercises"].add(ex["id"])
            for st in ex["sets"]:
                s["totalSets"] += 1
                if "weightKg" in st:
                    s["totalTonnageKg"] += st["reps"] * st["weightKg"]

    stats = {}
    for month, s in sorted(stats_by_month.items()):
        stats[month] = {
            "workouts": s["workouts"],
            "exerciseSessions": s["exerciseSessions"],
            "totalSets": s["totalSets"],
            "totalTonnageKg": round(s["totalTonnageKg"], 1),
            "uniqueExercises": len(s["uniqueExercises"]),
        }

    out = {
        "workouts": workouts,
        "exercises": exercises,
        "programs": programs,
        "stats": stats,
        "meta": {
            "source": "workouts_2026.xlsx",
            "rangeStart": workouts[0]["date"] if workouts else None,
            "rangeEnd": workouts[-1]["date"] if workouts else None,
            "workoutCount": len(workouts),
            "exerciseCount": len(exercises),
        },
    }

    DST.parent.mkdir(parents=True, exist_ok=True)
    DST.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK → {DST}")
    print(f"Workouts: {len(workouts)}")
    print(f"Unique exercises: {len(exercises)}")
    print(f"Programs: {len(programs)}")
    print(f"Monthly stats:")
    for m, s in stats.items():
        print(f"  {m}: {s['workouts']} трен., {s['totalSets']} подх., {s['totalTonnageKg']} кг")


if __name__ == "__main__":
    main()
